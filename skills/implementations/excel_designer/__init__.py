"""
Excel Designer Skill
Read, write, and style Excel (.xlsx) files using openpyxl.
"""
from __future__ import annotations

import io
from pathlib import Path
from typing import Any, Optional

from nexus.skills.registry import BaseSkill, SkillMeta

SKILL_META = SkillMeta(
    name="excel_designer",
    description="Read, create, and style Excel workbooks. "
                "Supports reading sheets, writing data tables, applying styles, "
                "creating charts, and saving as .xlsx",
    version="1.0.0",
    domains=["analysis", "operations", "creative"],
    triggers=["excel", "xlsx", "spreadsheet", "workbook", "表格", "試算表"],
    requires=["openpyxl"],
    is_local=True,
)


class Skill(BaseSkill):
    meta = SKILL_META

    async def run(
        self,
        operation: str,          # read | write | create | update | get_sheets
        file_path: str = "",
        sheet_name: str = "Sheet1",
        data: Optional[list[list[Any]]] = None,
        headers: Optional[list[str]] = None,
        start_row: int = 1,
        start_col: int = 1,
        style: Optional[dict] = None,
        output_path: Optional[str] = None,
        **kwargs,
    ) -> Any:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            return {"error": "openpyxl not installed. Run: pip install openpyxl"}

        op = operation.lower()

        # ── Read ──────────────────────────────────────────────────────────────
        if op == "read":
            if not file_path or not Path(file_path).exists():
                return {"error": f"File not found: {file_path}"}
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append(list(row))
            wb.close()
            return {
                "sheet": ws.title,
                "rows":  rows,
                "shape": (ws.max_row, ws.max_column),
            }

        # ── Get sheet names ───────────────────────────────────────────────────
        elif op == "get_sheets":
            if not file_path or not Path(file_path).exists():
                return {"error": f"File not found: {file_path}"}
            wb = openpyxl.load_workbook(file_path, read_only=True)
            sheets = wb.sheetnames
            wb.close()
            return {"sheets": sheets}

        # ── Create / Write ────────────────────────────────────────────────────
        elif op in ("create", "write"):
            # Load existing or create new workbook
            if file_path and Path(file_path).exists() and op == "write":
                wb = openpyxl.load_workbook(file_path)
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                else:
                    ws = wb.create_sheet(sheet_name)
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = sheet_name

            # Write headers with bold style
            if headers:
                for col_idx, header in enumerate(headers, start=start_col):
                    cell = ws.cell(row=start_row, column=col_idx, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(
                        start_color="366092", end_color="366092", fill_type="solid"
                    )
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.alignment = Alignment(horizontal="center")
                start_row += 1

            # Write data rows
            if data:
                for row_idx, row in enumerate(data, start=start_row):
                    for col_idx, value in enumerate(
                        (row if isinstance(row, (list, tuple)) else [row]),
                        start=start_col,
                    ):
                        ws.cell(row=row_idx, column=col_idx, value=value)

            # Auto-fit column widths
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

            save_path = output_path or file_path or "output.xlsx"
            Path(save_path).parent.mkdir(parents=True, exist_ok=True)
            wb.save(save_path)
            wb.close()
            return {"saved": save_path, "rows_written": len(data or []), "sheet": sheet_name}

        # ── Update single cell ────────────────────────────────────────────────
        elif op == "update":
            row    = kwargs.get("row", 1)
            col    = kwargs.get("col", 1)
            value  = kwargs.get("value")
            if not file_path:
                return {"error": "file_path required for update"}
            wb = openpyxl.load_workbook(file_path)
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
            ws.cell(row=row, column=col, value=value)
            wb.save(file_path)
            wb.close()
            return {"updated": f"({row},{col}) = {value}"}

        return {"error": f"Unknown operation: {operation}"}
